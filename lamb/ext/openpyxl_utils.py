from __future__ import annotations

import logging
from typing import Dict, List, Union, Callable, Optional, Generator

# Lamb Framework
from lamb.exc import (
    ApiError,
    InvalidParamTypeError,
    InvalidParamValueError,
    InvalidBodyStructureError,
)
from lamb.utils.core import compact

import openpyxl
from lamb.utils.core import lazy
from openpyxl.workbook import Workbook as OpenpyxlWorkbook
from openpyxl.cell.cell import Cell as OpenpyxlCell
from openpyxl.worksheet.worksheet import Worksheet as OpenpyxlWorksheet

__all__ = ["Worksheet", "Workbook", "Cell", "Row", "Column"]


logger = logging.getLogger(__name__)


# Column names based wrapper around Excel worksheet
# - use zero-based indexes instead of default openpyxl mode


class Workbook(object):
    def __init__(self, filename: Optional[str] = None, create_columns: bool = True, read_only: bool = False):
        if filename is not None:
            self._workbook = openpyxl.load_workbook(filename, data_only=True, read_only=read_only)
        else:
            self._workbook = OpenpyxlWorkbook()
        self._create_columns = create_columns
        self._filename = filename

    @property
    def worksheets(self) -> List["Worksheet"]:
        return [Worksheet(wrapped_worskheet=w, create_columns=self._create_columns) for w in self._workbook.worksheets]

    @property
    def worksheets_dict(self) -> Dict[str, "Worksheet"]:
        return {
            w.title: Worksheet(wrapped_worskheet=w, create_columns=self._create_columns)
            for w in self._workbook.worksheets
        }

    def remove_sheet(self, sheet: Union[str, "Worksheet"]):
        if isinstance(sheet, str):
            sheet = self.worksheets_dict[sheet]
        self._workbook.remove(sheet.openpyxl_worksheet)

    def create_sheet(self, title: str = None, index: int = None) -> "Worksheet":
        return Worksheet(self._workbook.create_sheet(title=title, index=index), create_columns=self._create_columns)

    # excel general
    @property
    def filename(self) -> str:
        return self._filename

    @property
    def openpyxl_workbook(self) -> OpenpyxlWorkbook:
        return self._workbook

    # iterators
    def iter_rows(self) -> Generator[Generator["Cell", None, None], None, None]:
        max_row = self._worksheet.max_row - 1
        for row in range(0, max_row):
            yield self.cells(row=row)

    def save(self, outputfile_path: str = None):
        if outputfile_path is None:
            outputfile_path = self._filename
        self._workbook.save(outputfile_path)

    def clean_sheet_empty_columns(self, sheet_name: str):
        # check and extract info
        if sheet_name not in self.worksheets_dict:
            raise InvalidParamValueError(f"Did not found sheet {sheet_name} in excel file")
        index = list(self.worksheets_dict.keys()).index(sheet_name)

        # clean
        ws = self.worksheets_dict[sheet_name]
        columns: List[List[Optional[str]]] = []

        removed_count = 0
        for column in ws.columns:
            values: List[Optional[str]] = [cell.value for cell in column.cells]
            if len(compact(values)) > 0:
                columns.append(values)
            else:
                removed_count += 1

        # save new sheet
        self.remove_sheet(sheet_name)
        ws = self.create_sheet(title=sheet_name, index=index)
        for column_index, column in enumerate(columns):
            for row_index, value in enumerate(column):
                ws.cell(row_index, column_index).value = value

        logger.info(f"columns cleaning: {sheet_name} -> removed columns count {removed_count}")

    def clean_sheet_empty_rows(self, sheet_name: str):
        # check and extract info
        if sheet_name not in self.worksheets_dict:
            raise InvalidParamValueError(f"Did not found sheet {sheet_name} in excel file")
        index = list(self.worksheets_dict.keys()).index(sheet_name)

        # clean
        ws = self.worksheets_dict[sheet_name]
        rows: List[List[Optional[str]]] = []

        removed_count = 0
        for row in ws.rows:
            values: List[Optional[str]] = [cell.value for cell in row.cells]
            if len(compact(values)) > 0:
                rows.append(values)
            else:
                removed_count += 1

        # save new sheet
        self.remove_sheet(sheet_name)
        ws = self.create_sheet(sheet_name, index=index)
        for row_index, row in enumerate(rows):
            for column_index, value in enumerate(row):
                ws.cell(row_index, column_index).value = value

        logger.info(f"rows cleaning: {sheet_name} -> removed rows count {removed_count}")


class Worksheet(object):
    _wrapped_worksheet: OpenpyxlWorksheet
    _create_columns: bool

    def __init__(self, wrapped_worskheet, create_columns):
        self._wrapped_worksheet = wrapped_worskheet
        self._create_columns = create_columns

    @property
    def openpyxl_worksheet(self) -> OpenpyxlWorksheet:
        return self._wrapped_worksheet

    @lazy
    def headers(self) -> List[Optional[str]]:
        return [cell.typed_value(req_type=str, default=None) for cell in self.cells(0)]

    @property
    def rows(self) -> Generator["Row", None, None]:
        max_row = self._wrapped_worksheet.max_row
        for row_index in range(0, max_row):
            yield Row(worksheet=self, row_index=row_index)

    @property
    def columns(self) -> Generator["Column", None, None]:
        max_column = self._wrapped_worksheet.max_column
        for column_index in range(0, max_column):
            yield Column(worksheet=self, column_index=column_index)

    # cell access
    def cells(self, row: int) -> Generator["Cell", None, None]:
        max_column = self._wrapped_worksheet.max_column
        for column in range(1, max_column + 1):
            result_cell = Cell(self._wrapped_worksheet.cell(row=row + 1, column=column))
            yield result_cell

    def cell(self, row: int, column: Union[int, str]) -> "Cell":
        # check params
        if isinstance(column, str):
            column_index = self.column_index(column)
            if column_index is None:
                if self._create_columns:
                    max_column = self._wrapped_worksheet.max_column
                    self._wrapped_worksheet.cell(row=1, column=max_column + 1).value = column
                    column_index = max_column
                    lazy.invalidate(self, "headers")
                else:
                    raise InvalidParamValueError(f"Column with name={column} not exist")
        elif isinstance(column, int):
            column_index = column
        else:
            raise InvalidParamTypeError("Invalid value of column param")

        if not isinstance(row, int):
            raise InvalidParamTypeError("Invalid value of row param")
        else:
            row_index = row

        # extract value - openpyxl use 1-based indexs
        openpyxl_cell = self._wrapped_worksheet.cell(row=row_index + 1, column=column_index + 1)
        return Cell(cell=openpyxl_cell)

    # utilities
    def column_name(self, column_index) -> Optional[str]:
        _headers = self.headers
        if column_index < len(_headers):
            return _headers[column_index]
        else:
            return None

    def column_index(self, column_name: str) -> Optional[int]:
        _headers = self.headers
        if column_name in _headers:
            return _headers.index(column_name)
        else:
            return None


class Row(object):
    def __init__(self, worksheet: Worksheet, row_index: int):
        self._worksheet = worksheet
        self._row_index = row_index

    def __getitem__(self, column_name) -> "Cell":
        return self._worksheet.cell(row=self._row_index, column=column_name)

    def __setitem__(self, column_name, value):
        self._worksheet.cell(row=self._row_index, column=column_name).value = value

    @property
    def cells(self) -> Generator["Cell", None, None]:
        _row = next(
            self._worksheet.openpyxl_worksheet.iter_rows(
                min_row=self._row_index + 1,
                max_row=self._row_index + 1,
                min_col=1,
                max_col=self._worksheet.openpyxl_worksheet.max_column,
            )
        )
        for _cell in _row:
            yield Cell(_cell)


class Column(object):
    def __init__(self, worksheet: Worksheet, column_index: int):
        self._worksheet = worksheet
        self._column_index = column_index

    @property
    def cells(self) -> Generator["Cell", None, None]:
        max_row = self._worksheet.openpyxl_worksheet.max_row
        for row_index in range(0, max_row):
            yield self._worksheet.cell(row=row_index, column=self._column_index)


class Cell(object):
    def __init__(self, cell: OpenpyxlCell):
        self._cell = cell

    def typed_value(self, req_type: type, allow_none: bool = False, transform: Callable = None, **kwargs):
        def _type_convert(_result):
            if req_type is None:
                return _result
            if isinstance(_result, req_type):
                return _result
            try:
                _result = req_type(_result)
                return _result
            except (ValueError, TypeError) as _e:
                raise InvalidParamTypeError("Invalid data type for value in cell") from _e

        try:
            # get and normalize value
            result = self._cell.value
            if len(str(result)) == 0:
                result = None

            # check none
            if result is None:
                if allow_none:
                    return None
                else:
                    raise InvalidParamTypeError("Invalid data type for value in cell")

            # apply type convert
            result = _type_convert(result)

            # apply transform
            if transform is not None:
                return transform(result)

            return result
        except Exception as e:
            if "default" in kwargs.keys():
                return kwargs["default"]
            elif isinstance(e, ApiError):
                raise
            else:
                raise InvalidBodyStructureError("Failed to parse cell value") from e

    @property
    def openpyxl_cell(self) -> OpenpyxlCell:
        return self._cell

    @property
    def value(self):
        return self.typed_value(req_type=str, allow_none=True)

    @value.setter
    def value(self, value):
        self._cell.value = value
